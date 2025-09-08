#!/usr/bin/env python3
"""
COA Migration Excel Report Generator
Generates Excel reports for Chart of Accounts migration validation

Dependencies:
    pip install openpyxl pandas
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import json
from decimal import Decimal

class COAMigrationExcelGenerator:
    """Generate Excel reports for COA migration"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.styles = self._create_styles()
    
    def _create_styles(self):
        """Create Excel styling"""
        return {
            'header': Font(bold=True, color="FFFFFF"),
            'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'subheader': Font(bold=True, size=12),
            'subheader_fill': PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            'currency': '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)',
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'success_fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'error_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'warning_fill': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        }
    
    def create_migration_summary_sheet(self, migration_data: dict):
        """Create migration summary overview sheet"""
        ws = self.workbook.active
        ws.title = "Migration Summary"
        
        # Title
        ws['A1'] = "OSUSAPPS - Chart of Accounts Migration Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Migration details
        row = 3
        details = [
            ("Migration Date:", migration_data.get('migration_date', datetime.now().strftime('%Y-%m-%d'))),
            ("Target System:", "OSUSAPPS (Odoo 17)"),
            ("Source Systems:", ", ".join(migration_data.get('source_systems', []))),
            ("Total Accounts Migrated:", migration_data.get('total_accounts_migrated', 0)),
            ("Migration Status:", migration_data.get('migration_status', 'Unknown'))
        ]
        
        for label, value in details:
            ws.cell(row=row, column=1, value=label).font = self.styles['subheader']
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Balance validation summary
        row += 2
        ws.cell(row=row, column=1, value="Balance Validation Summary").font = Font(bold=True, size=14)
        row += 1
        
        # Headers
        headers = ["System", "Total Debits", "Total Credits", "Difference", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center']
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[chr(65 + col - 1)].width = 20
    
    def create_trial_balance_sheet(self, tb_data: list, sheet_name: str):
        """Create trial balance sheet"""
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Convert to DataFrame
        df = pd.DataFrame(tb_data)
        
        # Add title
        ws['A1'] = f"{sheet_name} - Trial Balance"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Add data starting from row 3
        start_row = 3
        
        # Headers
        headers = ["Account Code", "Account Name", "Account Type", "Debit Balance", "Credit Balance", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center']
        
        # Data rows
        row = start_row + 1
        total_debits = Decimal('0.00')
        total_credits = Decimal('0.00')
        
        for record in tb_data:
            ws.cell(row=row, column=1, value=record.get('Account_Code', ''))
            ws.cell(row=row, column=2, value=record.get('Account_Name', ''))
            ws.cell(row=row, column=3, value=record.get('Account_Type', ''))
            
            # Debit balance
            debit = float(record.get('Debit_Balance', 0))
            debit_cell = ws.cell(row=row, column=4, value=debit)
            debit_cell.number_format = self.styles['currency']
            total_debits += Decimal(str(debit))
            
            # Credit balance  
            credit = float(record.get('Credit_Balance', 0))
            credit_cell = ws.cell(row=row, column=5, value=credit)
            credit_cell.number_format = self.styles['currency']
            total_credits += Decimal(str(credit))
            
            ws.cell(row=row, column=6, value=record.get('Notes', ''))
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.styles['border']
            
            row += 1
        
        # Totals row
        row += 1
        ws.cell(row=row, column=2, value="TOTALS").font = self.styles['subheader']
        
        total_debit_cell = ws.cell(row=row, column=4, value=float(total_debits))
        total_debit_cell.number_format = self.styles['currency']
        total_debit_cell.font = self.styles['subheader']
        total_debit_cell.fill = self.styles['subheader_fill']
        
        total_credit_cell = ws.cell(row=row, column=5, value=float(total_credits))
        total_credit_cell.number_format = self.styles['currency']
        total_credit_cell.font = self.styles['subheader']
        total_credit_cell.fill = self.styles['subheader_fill']
        
        # Balance check
        difference = total_debits - total_credits
        balance_status = "✓ BALANCED" if abs(difference) < Decimal('0.01') else f"✗ OUT OF BALANCE (${difference})"
        balance_cell = ws.cell(row=row, column=6, value=balance_status)
        balance_cell.font = self.styles['subheader']
        balance_cell.fill = self.styles['success_fill'] if "BALANCED" in balance_status else self.styles['error_fill']
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[chr(65 + col - 1)].width = 18
    
    def create_account_mapping_sheet(self, mapping_data: list):
        """Create account mapping sheet"""
        ws = self.workbook.create_sheet(title="Account Mappings")
        
        # Title
        ws['A1'] = "Chart of Accounts - Account Mappings"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Headers
        start_row = 3
        headers = ["Source System", "Legacy Code", "Legacy Name", "New Code", "New Name", "Odoo Account Type", "Category", "Notes"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center']
        
        # Data rows
        row = start_row + 1
        for mapping in mapping_data:
            ws.cell(row=row, column=1, value=mapping.get('source_system', ''))
            ws.cell(row=row, column=2, value=mapping.get('legacy_code', ''))
            ws.cell(row=row, column=3, value=mapping.get('legacy_name', ''))
            ws.cell(row=row, column=4, value=mapping.get('new_code', ''))
            ws.cell(row=row, column=5, value=mapping.get('new_name', ''))
            ws.cell(row=row, column=6, value=mapping.get('odoo_account_type', ''))
            ws.cell(row=row, column=7, value=mapping.get('account_category', ''))
            ws.cell(row=row, column=8, value=mapping.get('notes', ''))
            
            # Apply borders and alternate row coloring
            fill = self.styles['subheader_fill'] if row % 2 == 0 else None
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = self.styles['border']
                if fill:
                    cell.fill = fill
            
            row += 1
        
        # Auto-adjust column widths
        column_widths = [15, 12, 25, 12, 25, 20, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(65 + col - 1)].width = width
    
    def create_balance_sheet_comparison(self, pre_migration: dict, post_migration: dict):
        """Create balance sheet comparison"""
        ws = self.workbook.create_sheet(title="Balance Sheet Comparison")
        
        # Title
        ws['A1'] = "Balance Sheet - Pre vs Post Migration"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Headers
        start_row = 3
        headers = ["Account Category", "Pre-Migration", "Post-Migration", "Variance", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.border = self.styles['border']
            cell.alignment = self.styles['center']
        
        # Sample balance sheet data structure
        balance_categories = [
            "Total Assets",
            "Current Assets", 
            "Fixed Assets",
            "Total Liabilities",
            "Current Liabilities",
            "Long-term Liabilities",
            "Total Equity",
            "Retained Earnings"
        ]
        
        row = start_row + 1
        for category in balance_categories:
            ws.cell(row=row, column=1, value=category).font = self.styles['subheader']
            
            # Placeholder values - replace with actual data
            pre_value = 0.00
            post_value = 0.00
            variance = post_value - pre_value
            status = "✓" if variance == 0 else "⚠"
            
            # Pre-migration value
            pre_cell = ws.cell(row=row, column=2, value=pre_value)
            pre_cell.number_format = self.styles['currency']
            
            # Post-migration value
            post_cell = ws.cell(row=row, column=3, value=post_value)
            post_cell.number_format = self.styles['currency']
            
            # Variance
            var_cell = ws.cell(row=row, column=4, value=variance)
            var_cell.number_format = self.styles['currency']
            var_cell.fill = self.styles['success_fill'] if variance == 0 else self.styles['warning_fill']
            
            # Status
            status_cell = ws.cell(row=row, column=5, value=status)
            status_cell.alignment = self.styles['center']
            status_cell.fill = self.styles['success_fill'] if status == "✓" else self.styles['warning_fill']
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.styles['border']
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[chr(65 + col - 1)].width = 18
    
    def create_validation_results_sheet(self, validation_data: dict):
        """Create validation results sheet"""
        ws = self.workbook.create_sheet(title="Validation Results")
        
        # Title
        ws['A1'] = "Migration Validation Results"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Summary
        row = 3
        ws.cell(row=row, column=1, value="Validation Summary").font = Font(bold=True, size=12)
        row += 1
        
        summary_items = [
            ("Total Errors:", validation_data.get('error_count', 0)),
            ("Total Warnings:", validation_data.get('warning_count', 0)),
            ("Overall Status:", "PASS" if validation_data.get('error_count', 0) == 0 else "FAIL")
        ]
        
        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label).font = self.styles['subheader']
            cell = ws.cell(row=row, column=2, value=value)
            if label == "Overall Status:":
                cell.fill = self.styles['success_fill'] if value == "PASS" else self.styles['error_fill']
            row += 1
        
        # Errors section
        if validation_data.get('errors'):
            row += 2
            ws.cell(row=row, column=1, value="Validation Errors").font = Font(bold=True, size=12, color="FF0000")
            row += 1
            
            for error in validation_data['errors']:
                cell = ws.cell(row=row, column=1, value=f"✗ {error}")
                cell.fill = self.styles['error_fill']
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
        
        # Warnings section
        if validation_data.get('warnings'):
            row += 2
            ws.cell(row=row, column=1, value="Validation Warnings").font = Font(bold=True, size=12, color="FF8C00")
            row += 1
            
            for warning in validation_data['warnings']:
                cell = ws.cell(row=row, column=1, value=f"⚠ {warning}")
                cell.fill = self.styles['warning_fill']
                ws.merge_cells(f'A{row}:D{row}')
                row += 1
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 60
    
    def generate_complete_report(self, migration_report: dict, output_file: str = 'COA_Migration_Report.xlsx'):
        """Generate complete Excel migration report"""
        try:
            # Migration Summary (already created as active sheet)
            self.create_migration_summary_sheet(migration_report.get('migration_summary', {}))
            
            # Trial Balance sheets (if data available)
            balance_data = migration_report.get('balance_validation', {})
            
            # Create sample OSUS Properties trial balance
            osus_sample_data = [
                {'Account_Code': '1001', 'Account_Name': 'Cash - Operating', 'Account_Type': 'Asset', 'Debit_Balance': '25000.00', 'Credit_Balance': '0.00', 'Notes': 'Primary operating cash'},
                {'Account_Code': '1100', 'Account_Name': 'Accounts Receivable', 'Account_Type': 'Asset', 'Debit_Balance': '45000.00', 'Credit_Balance': '0.00', 'Notes': 'Customer receivables'},
                {'Account_Code': '2001', 'Account_Name': 'Accounts Payable', 'Account_Type': 'Liability', 'Debit_Balance': '0.00', 'Credit_Balance': '25000.00', 'Notes': 'Vendor payables'},
                {'Account_Code': '3001', 'Account_Name': 'Common Stock', 'Account_Type': 'Equity', 'Debit_Balance': '0.00', 'Credit_Balance': '45000.00', 'Notes': 'Issued shares'}
            ]
            self.create_trial_balance_sheet(osus_sample_data, "OSUS Properties TB")
            
            # Account Mappings
            if migration_report.get('account_mappings'):
                self.create_account_mapping_sheet(migration_report['account_mappings'])
            
            # Balance Sheet Comparison
            self.create_balance_sheet_comparison({}, {})
            
            # Validation Results
            if migration_report.get('validation_results'):
                self.create_validation_results_sheet(migration_report['validation_results'])
            
            # Save the workbook
            self.workbook.save(output_file)
            print(f"✓ Excel report generated: {output_file}")
            return True
            
        except Exception as e:
            print(f"✗ Error generating Excel report: {str(e)}")
            return False

def main():
    """Generate sample migration Excel report"""
    # Sample migration data
    sample_report = {
        'migration_summary': {
            'migration_date': '2025-09-07',
            'source_systems': ['OSUS Properties', 'Legacy System'],
            'target_system': 'OSUSAPPS (Odoo 17)',
            'total_accounts_migrated': 27,
            'migration_status': 'Completed Successfully'
        },
        'balance_validation': {
            'osus_properties': {
                'total_debits': 265000.00,
                'total_credits': 265000.00,
                'account_count': 20,
                'is_balanced': True
            },
            'legacy_system': {
                'total_debits': 188000.00,
                'total_credits': 188000.00,
                'account_count': 7,
                'is_balanced': True
            }
        },
        'account_mappings': [
            {
                'source_system': 'OSUS_Properties',
                'legacy_code': '1001',
                'legacy_name': 'Cash - Operating',
                'new_code': '101001',
                'new_name': 'Cash - Operating Account',
                'odoo_account_type': 'asset_cash',
                'account_category': 'Asset',
                'notes': 'Migrated from OSUS Properties'
            }
        ],
        'validation_results': {
            'errors': [],
            'warnings': ['Account 1500 has both debit and credit balances'],
            'error_count': 0,
            'warning_count': 1
        }
    }
    
    # Generate report
    generator = COAMigrationExcelGenerator()
    generator.generate_complete_report(sample_report, 'OSUSAPPS_COA_Migration_Report.xlsx')

if __name__ == "__main__":
    main()
